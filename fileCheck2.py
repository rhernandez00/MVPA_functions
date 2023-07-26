# -*- coding: utf-8 -*-
"""
Created on Mon Jun 12 20:45:57 2023

@author: Raul
"""

import tkinter as tk
from tkinter import filedialog, ttk
import pandas as pd
from openpyxl import load_workbook

import os
import re
from itertools import product


def evaluateSheet(tablePath,sheetName):
    data = pd.read_excel(tablePath,sheet_name='Variables')
    variablesDict = {}
    for index,row in data.iterrows():
        variablesDict[row.varName] = transform(row.varValues,row.varType,row.paddedBy)
    dataS = pd.read_excel(tablePath,sheet_name=sheetName)
    d = pd.DataFrame()
    for indx,row in dataS.iterrows():
        s = dataS.iloc[indx]['path']  + '\\' + dataS.iloc[indx]['baseName']
        newList = explodeString(s,variablesDict)
        for fileName in newList:
            if os.path.isfile(fileName):
                file_stat = os.stat(fileName)
                fileSize = round(file_stat.st_size / (1024 * 1024),4)
            else:
                fileSize = 'NA'
            fileRow = pd.Series({
                'fileName': fileName,
                'exist': os.path.isfile(fileName),
                'size': fileSize,
            })
            d = d.append(fileRow, ignore_index=True)
    d = d.sort_values(['exist','fileName'],ascending=[True,True])
    return d

def transform(string, varType, paddedBy):
    result = []
    if varType == 'number':
        parts = string.split(',')  # Split the string by commas

        for part in parts:
            part = part.strip()  # Remove leading and trailing whitespace
            if ':' in part:
                start, end = part.split(':')  # Split the part by colon if it contains a range
                result.extend([str(i).zfill(paddedBy) for i in range(int(start), int(end) + 1)])  # Add the range of numbers to the result list, as zero-padded strings
            else:
                result.append(str(int(part)).zfill(paddedBy))  # Add the single number to the result list, as a zero-padded string
    elif varType == 'string':
        parts = string.split(',')  # Split the string by commas
        for part in parts:
            part = part.strip()  # Remove leading and trailing whitespace
            result.append(part)  # Add each element of the list
    else:
        raise ValueError("Error: Wrong varType: " + varType)

    return result


def explodeString(s,variablesDict):
    #Finds the variables in the string
    variableNames = re.findall(r'\[(.*?)\]', s)

    #Gets all the elements of the list for the given variable
    variableValues = product(*(variablesDict[name] for name in variableNames))

    # For each combination, substitute the variables in s with the corresponding values
    resList = []
    for values in variableValues:
        resList.append(s)
        for name, value in zip(variableNames, values):
            resList[-1] = resList[-1].replace('['+name+']', str(value))
    correctedList = []
    for n in resList:
        correctedList.append(n.replace('\\\\', '\\'))
    return correctedList

def load_xlsx():
    global tablePath
    tablePath = filedialog.askopenfilename(filetypes=(("Excel files", "*.xlsx"), ("All files", "*.*")))
    wb = load_workbook(tablePath)
    sheets = wb.sheetnames

    for i, sheet in enumerate(sheets):
        if sheet != 'Variables':
            btn = tk.Button(root, text=sheet, command=lambda sheet=sheet: show_table(sheet))
            btn.grid(row=i+2, column=0, sticky='w')

def show_table(sheetName):
    df = evaluateSheet(tablePath, sheetName)
    
    # Determine column widths based on the longest element in each column
    column_widths = {}
    for column in df.columns:
        max_length = max(df[column].astype(str).map(len).max(), len(column))
        column_widths[column] = max_length
    
    tree = ttk.Treeview(root, columns=df.columns.values.tolist())
    for column in df.columns:
        #print(column)
        tree.heading(column, text=column)
        tree.column(column, width=column_widths[column] * 10)

    for i in df.index:
        values = df.loc[i].values.tolist()
        if 'exist' in df.columns:
            if df['exist'][i] == 1.0:
                values[0] = 'true'
            else:
                values[0] = 'missing'
        tree.insert('', 'end', values=values)

    tree.grid(row=0, column=0, rowspan=len(df.index))

root = tk.Tk()
load_button = tk.Button(root, text="Load xlsx file", command=load_xlsx)
load_button.grid(row=0, column=1, sticky='w')

root.mainloop()